import os
import sys
import shutil
import time
import json
import threading
import hashlib
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple, List, Dict

from PySide6 import QtCore, QtGui, QtWidgets
from openpyxl import Workbook
from openpyxl.styles import Font

# Optional metadata libs
try:
    from PIL import Image
    PIL_OK = True
except Exception:
    PIL_OK = False

try:
    from mutagen import File as MutagenFile
    MUTAGEN_OK = True
except Exception:
    MUTAGEN_OK = False

# ---------------------------
# Config
# ---------------------------
IMAGE_EXTS = {
    ".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".tif", ".webp", ".heic", ".heif"
}
VIDEO_EXTS = {
    ".mp4", ".mov", ".avi", ".mkv", ".wmv", ".flv", ".webm", ".m4v", ".3gp"
}
SUPPORTED_EXTS = IMAGE_EXTS | VIDEO_EXTS

APP_TITLE = "Ragilmalik's Media Sorter -- Image & Video Organizer"
INDEX_FILENAME = ".media_sorter_index.json"

# Month names (Indonesian)
ID_MONTHS = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April", 5: "Mei", 6: "Juni",
    7: "Juli", 8: "Agustus", 9: "September", 10: "Oktober", 11: "November", 12: "Desember"
}
ID_MONTHS_ABBR = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "Mei", 6: "Jun",
    7: "Jul", 8: "Agu", 9: "Sep", 10: "Okt", 11: "Nov", 12: "Des"
}

# ---------------------------
# Modern Themes (pure black/white core, accents only on outlines)
# ---------------------------
DARK_QSS = """
* { font-family: Segoe UI, Inter, Arial; }
QWidget { background: #000000; color: #FFFFFF; }
QGroupBox { border: 1px solid #0EA5A4; border-radius: 12px; margin-top: 16px; }
QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 6px; } /* default text color */
QLineEdit, QTextEdit, QComboBox, QListView, QTreeView, QTableView {
    background: #0A0A0A; border: 1px solid #0EA5A4; border-radius: 10px; color: #FFFFFF; padding: 8px 12px;
}
QComboBox::drop-down { width: 26px; border-left: 1px solid #0EA5A4; }
QComboBox QAbstractItemView { padding: 6px; border: 1px solid #1F2937; background: #0A0A0A; color: #FFFFFF; }
QPushButton {
    background: #0B0B0B; color: #FFFFFF; border: 1px solid #0EA5A4; border-radius: 12px; padding: 10px 14px; font-weight: 600;
}
QPushButton:hover { background: #FFFFFF; color: #000000; } /* opposite on hover */
QPushButton:disabled { color: #6B7280; border-color: #1F2937; }
QCheckBox, QLabel { color: #FFFFFF; }
QProgressBar { background: #0A0A0A; border: 1px solid #1F2937; border-radius: 8px; text-align: center; color: #FFFFFF; }
QProgressBar::chunk { background-color: #FFFFFF; } /* neutral fill to honor 'outline-only' accent rule */
QTableWidget { gridline-color: #0EA5A4; } /* cyan outline for table grid */
QHeaderView::section { background: #0B0B0B; color: #FFFFFF; border: 0px; padding: 8px 10px; }
"""

LIGHT_QSS = """
* { font-family: Segoe UI, Inter, Arial; }
QWidget { background: #FFFFFF; color: #000000; }
QGroupBox { border: 1px solid #2563EB; border-radius: 12px; margin-top: 16px; }
QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 6px; } /* default text color */
QLineEdit, QTextEdit, QComboBox, QListView, QTreeView, QTableView {
    background: #FAFAFA; border: 1px solid #2563EB; border-radius: 10px; color: #000000; padding: 8px 12px;
}
QComboBox::drop-down { width: 26px; border-left: 1px solid #2563EB; }
QComboBox QAbstractItemView { padding: 6px; border: 1px solid #E5E7EB; background: #FFFFFF; color: #000000; }
QPushButton {
    background: #F3F4F6; color: #000000; border: 1px solid #2563EB; border-radius: 12px; padding: 10px 14px; font-weight: 600;
}
QPushButton:hover { background: #000000; color: #FFFFFF; } /* opposite on hover */
QPushButton:disabled { color: #9AA0A6; border-color: #E5E7EB; }
QCheckBox, QLabel { color: #000000; }
QProgressBar { background: #F3F4F6; border: 1px solid #E5E7EB; border-radius: 8px; text-align: center; color: #000000; }
QProgressBar::chunk { background-color: #000000; } /* neutral fill */
QTableWidget { gridline-color: #2563EB; } /* blue outline for table grid */
QHeaderView::section { background: #EEF2FF; color: #000000; border: 0px; padding: 8px 10px; }
"""

# ---------------------------
# Helpers
# ---------------------------

def is_media_file(path: Path) -> bool:
    return path.is_file() and path.suffix.lower() in SUPPORTED_EXTS

def to_long_path(path: Path) -> str:
    """Return a Windows long-path-safe string; on other OS returns str(path)."""
    s = str(path)
    if os.name == "nt":
        s = os.path.abspath(s)
        if s.startswith("\\\\") and not s.startswith("\\\\?\\"):
            # UNC path -> \\?\UNC\server\share\...
            s = "\\\\?\\UNC\\" + s[2:]
        elif not s.startswith("\\\\?\\"):
            # Drive path -> \\?\C:\...
            s = "\\\\?\\" + s
    return s

def ensure_unique_path(dest_path: Path) -> Path:
    """If dest_path exists, append (1), (2), ... before the suffix."""
    if not dest_path.exists():
        return dest_path
    stem = dest_path.stem
    suffix = dest_path.suffix
    parent = dest_path.parent
    i = 1
    while True:
        candidate = parent / f"{stem} ({i}){suffix}"
        if not candidate.exists():
            return candidate
        i += 1

# ---- EXIF & metadata ----

def parse_exif_datetime(dt_str: str) -> Optional[datetime]:
    for fmt in ("%Y:%m:%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y:%m:%d %H:%M:%S%z"):
        try:
            return datetime.strptime(dt_str.strip(), fmt).replace(tzinfo=None)
        except Exception:
            continue
    return None

def read_image_metadata(p: Path) -> Dict[str, Optional[object]]:
    md: Dict[str, Optional[object]] = {"width": None, "height": None, "camera_make": None, "camera_model": None,
          "lens": None, "gps_lat": None, "gps_lon": None, "date_taken": None}
    if not PIL_OK:
        return md
    try:
        with Image.open(p) as img:
            md["width"], md["height"] = img.size
            try:
                exif = img.getexif() or {}
                for tag in (36867, 36868, 306):  # DateTimeOriginal, DateTimeDigitized, DateTime
                    if tag in exif:
                        dt = parse_exif_datetime(str(exif.get(tag)))
                        if dt:
                            md["date_taken"] = dt
                            break
                md["camera_make"] = exif.get(271)
                md["camera_model"] = exif.get(272)
                md["lens"] = exif.get(42036) or exif.get(42035)
            except Exception:
                pass
    except Exception:
        pass
    return md

def parse_mp4_mov_datetime(tags: dict) -> Optional[datetime]:
    candidates = ["\xa9day", "creation_time", "com.apple.quicktime.creationdate"]
    for k in candidates:
        v = tags.get(k)
        if not v:
            continue
        if isinstance(v, list):
            v = v[0]
        v = str(v)
        for fmt in ("%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%d %H:%M:%S%z", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(v.strip(), fmt).replace(tzinfo=None)
            except Exception:
                continue
    return None

def read_video_metadata(p: Path) -> Dict[str, Optional[object]]:
    md: Dict[str, Optional[object]] = {"duration": None, "date_taken": None}
    if not MUTAGEN_OK:
        return md
    try:
        f = MutagenFile(str(p))
        if f is None:
            return md
        try:
            md["duration"] = getattr(f.info, "length", None)
        except Exception:
            pass
        tags = {}
        try:
            if hasattr(f, "tags") and f.tags is not None:
                tags = {str(k): v for k, v in f.tags.items()}
        except Exception:
            pass
        dt = parse_mp4_mov_datetime(tags) if tags else None
        if dt:
            md["date_taken"] = dt
    except Exception:
        pass
    return md

def safe_get_creation_dt(p: Path) -> datetime:
    """Filesystem times: creation (if any) else modified."""
    try:
        ctime = p.stat().st_ctime
    except Exception:
        ctime = 0
    try:
        mtime = p.stat().st_mtime
    except Exception:
        mtime = time.time()
    ts = ctime if ctime and ctime > 0 else mtime
    return datetime.fromtimestamp(ts)

def best_creation_datetime(p: Path) -> Tuple[datetime, Dict[str, Optional[object]]]:
    """Return best guess of date taken and extra metadata."""
    ext = p.suffix.lower()
    base_meta: Dict[str, Optional[object]] = {
        "width": None, "height": None, "camera_make": None, "camera_model": None,
        "lens": None, "gps_lat": None, "gps_lon": None, "duration": None,
    }
    if ext in IMAGE_EXTS:
        imd = read_image_metadata(p)
        base_meta.update({k: imd.get(k) for k in ("width", "height", "camera_make", "camera_model", "lens")})
        dt = imd.get("date_taken") or safe_get_creation_dt(p)
        return dt, base_meta
    if ext in VIDEO_EXTS:
        vmd = read_video_metadata(p)
        base_meta["duration"] = vmd.get("duration")
        dt = vmd.get("date_taken") or safe_get_creation_dt(p)
        return dt, base_meta
    return safe_get_creation_dt(p), base_meta

# ---- MD5 for exact duplicates ----

def file_md5(path: Path, chunk_size: int = 2 * 1024 * 1024) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()

# ---------------------------
# Worker
# ---------------------------
class SortWorker(QtCore.QThread):
    progress = QtCore.Signal(int, int)  # processed, total
    status = QtCore.Signal(str)
    finished_success = QtCore.Signal(dict)
    error = QtCore.Signal(str)

    def __init__(
        self,
        src_folder: str,
        dst_folder: str,
        recursive: bool,
        rename: bool,
        month_lang: str,
        template_key: str,
        simulate: bool,
        resume_skip: bool,
        dup_level: str,
        dup_action: str,
        dup_folder: str,
        parent=None,
    ):
        super().__init__(parent)
        self.src_folder = Path(src_folder)
        self.dst_folder = Path(dst_folder)
        self.recursive = recursive
        self.rename = rename
        self.month_lang = month_lang
        self.template_key = template_key
        self.simulate = simulate
        self.resume_skip = resume_skip
        self.dup_level = dup_level  # 'off' | 'exact'
        self.dup_action = dup_action  # 'skip' | 'keep' | 'folder'
        self.dup_folder = dup_folder.strip() or "Duplicates"

        self._stop = False
        self._pause_event = threading.Event()
        self._pause_event.set()
        self._rows_lock = threading.Lock()
        self._count_lock = threading.Lock()

        self.rows: List[Dict[str, object]] = []
        self.count_images = 0
        self.count_videos = 0
        self.count_renamed = 0
        self.total_files = 0
        self.workers_used = 1  # set by auto-tune

        self._index: Dict[str, Dict[str, object]] = {}
        self._index_path = self.dst_folder / INDEX_FILENAME
        self._md5_map: Dict[str, List[str]] = {}

    # ---- lifecycle ----
    def run(self):
        try:
            if not self.src_folder.exists():
                self.error.emit("Source folder does not exist.")
                return
            if not self.dst_folder.exists():
                self.dst_folder.mkdir(parents=True, exist_ok=True)

            self._load_index()
            self._rebuild_md5_from_index()

            files = self._collect_files()
            self.total_files = len(files)
            if self.total_files == 0:
                self.status.emit("No media files found.")
                self.progress.emit(0, 0)
                log_path = self._write_excel_log(simulation=self.simulate)
                summary = {
                    "total": 0, "images": 0, "videos": 0, "renamed": 0,
                    "log_path": str(log_path) if log_path else "", "simulate": self.simulate,
                }
                self.finished_success.emit(summary)
                return

            self.workers_used = self._auto_workers(files)
            self.status.emit(f"Found {self.total_files} media files. Using {self.workers_used} workers (auto-tuned).")
            self.progress.emit(0, self.total_files)

            processed = 0
            if self.workers_used == 1:
                for f in files:
                    if self._stop: break
                    self._process_one(f)
                    processed += 1
                    self.progress.emit(processed, self.total_files)
            else:
                with ThreadPoolExecutor(max_workers=self.workers_used) as ex:
                    futures = [ex.submit(self._process_one, f) for f in files]
                    for fut in as_completed(futures):
                        _ = fut.result()
                        processed += 1
                        self.progress.emit(processed, self.total_files)

            log_path = self._write_excel_log(simulation=self.simulate)
            if not self.simulate:
                self._save_index()

            summary = {
                "total": processed,
                "images": self.count_images,
                "videos": self.count_videos,
                "renamed": self.count_renamed,
                "log_path": str(log_path) if log_path else "",
                "simulate": self.simulate,
            }
            self.finished_success.emit(summary)
        except Exception as e:
            self.error.emit(str(e))

    def stop(self):
        self._stop = True

    def pause(self):
        self._pause_event.clear()

    def resume(self):
        self._pause_event.set()

    # ---- auto-tune workers ----
    def _auto_workers(self, files: List[Path]) -> int:
        base = min(6, max(2, (os.cpu_count() or 4)))
        try:
            sample = files[: min(60, len(files))]
            sizes = []
            for p in sample:
                try: sizes.append(p.stat().st_size)
                except Exception: pass
            if not sizes: return base
            sizes.sort()
            mid = len(sizes) // 2
            median = sizes[mid] if len(sizes) % 2 == 1 else (sizes[mid - 1] + sizes[mid]) / 2.0
            if median < 1 * 1024 * 1024:   # <1MB
                workers = min(6, base + 2)
            elif median < 8 * 1024 * 1024: # <8MB
                workers = base
            else:
                workers = max(2, base - 1)
            if str(self.src_folder).startswith("\\\\") or str(self.dst_folder).startswith("\\\\"):
                workers = min(workers, 3)
            return max(1, workers)
        except Exception:
            return base

    # ---- file enumeration ----
    def _collect_files(self) -> List[Path]:
        files: List[Path] = []
        if self.recursive:
            for root, _, filenames in os.walk(self.src_folder):
                for name in filenames:
                    p = Path(root) / name
                    if is_media_file(p): files.append(p)
        else:
            for p in self.src_folder.iterdir():
                if is_media_file(p): files.append(p)
        return files

    # ---- path building ----
    def _month_name(self, dt: datetime) -> str:
        if getattr(self, "month_lang", "en") == "id":
            return ID_MONTHS.get(dt.month, dt.strftime("%B"))
        return dt.strftime("%B")

    def _month_abbr(self, dt: datetime) -> str:
        if getattr(self, "month_lang", "en") == "id":
            return ID_MONTHS_ABBR.get(dt.month, dt.strftime("%b"))
        return dt.strftime("%b")

    def _build_dest_dir(self, dt: datetime) -> Path:
        # 'ymd_name' -> YYYY/MonthName/DD
        # 'ymd_mm'   -> YYYY/MM/DD
        # 'ymd_mon'  -> YYYY/Mon/DD
        yyyy = dt.strftime("%Y"); mm = dt.strftime("%m"); dd = dt.strftime("%d")
        if self.template_key == "ymd_mm":
            parts = [yyyy, mm, dd]
        elif self.template_key == "ymd_mon":
            parts = [yyyy, self._month_abbr(dt), dd]
        else:  # default 'ymd_name'
            parts = [yyyy, self._month_name(dt), dd]
        dest = self.dst_folder
        for seg in parts: dest = dest / seg
        return dest

    # ---- dedupe index ----
    def _rebuild_md5_from_index(self):
        self._md5_map.clear()
        for rec in self._index.values():
            md5 = rec.get("md5")
            if md5:
                self._md5_map.setdefault(str(md5), []).append(rec.get("dest", ""))

    def _detect_exact_duplicate(self, src_path: Path) -> Tuple[bool, str, Optional[str]]:
        if self.dup_level != "exact":
            return False, "", None
        try:
            md5_hex = file_md5(src_path)
            hits = self._md5_map.get(md5_hex)
            if hits:
                return True, hits[0], md5_hex
            return False, "", md5_hex
        except Exception:
            return False, "", None

    def _update_index(self, src_path: Path, size: int, mtime: float, dest_path: Path, md5_hex: Optional[str]):
        key = str(src_path.resolve())
        rec: Dict[str, object] = {"size": int(size), "mtime": float(mtime), "dest": str(dest_path)}
        if md5_hex:
            rec["md5"] = md5_hex
            self._md5_map.setdefault(md5_hex, []).append(str(dest_path))
        self._index[key] = rec

    def _should_skip_resume(self, src_path: Path, size: int, mtime: float) -> Tuple[bool, str]:
        if not self.resume_skip: return False, ""
        key = str(src_path.resolve()); rec = self._index.get(key)
        if not rec: return False, ""
        try:
            if int(rec.get("size", -1)) == size and abs(float(rec.get("mtime", 0)) - mtime) < 0.5:
                dest = Path(str(rec.get("dest", "")))
                if dest.exists(): return True, str(rec.get("dest", ""))
        except Exception: pass
        return False, ""

    # ---- per-file processing ----
    def _process_one(self, src_path: Path):
        if self._stop: return
        while not self._pause_event.is_set(): time.sleep(0.05)
        try:
            stat = src_path.stat()
            size = stat.st_size; mtime = stat.st_mtime

            cdt, extra = best_creation_datetime(src_path)
            dest_dir = self._build_dest_dir(cdt)
            dest_dir.mkdir(parents=True, exist_ok=True)

            new_name = src_path.name
            if self.rename:
                formatted_date = f"{cdt:%d-%m-%Y}"
                new_name = f"{formatted_date}-{src_path.name}"
            dest_path = ensure_unique_path(dest_dir / new_name)

            # Resume skip
            if self.resume_skip:
                skip, prev_dest = self._should_skip_resume(src_path, size, mtime)
                if skip:
                    self._append_row(src_path, dest_dir, new_name, cdt, extra, "Skipped", size, "")
                    self.status.emit(f"Skipped ▶ {src_path.name} (already copied)")
                    self._update_counts(src_path)
                    return

            # Exact duplicate handling
            is_dup, dup_path, md5_hex = self._detect_exact_duplicate(src_path)
            action = "Copied"; planned_dest = str(dest_path); do_copy = True; duplicate_of = ""
            if is_dup:
                duplicate_of = dup_path or ""
                if self.dup_action == "skip":
                    action = "Duplicate-Skipped"; do_copy = False
                elif self.dup_action == "folder":
                    # Place inside date folder: <dest_dir>/<dup_folder>/<file>
                    dup_dir = dest_dir / self.dup_folder
                    dup_dir.mkdir(parents=True, exist_ok=True)
                    dest_path = ensure_unique_path(dup_dir / new_name)
                    planned_dest = str(dest_path)
                    action = f"Duplicate→{self.dup_folder}"
                else:
                    action = "Duplicate-Kept"  # keep both

            if self.simulate:
                action = f"Simulated-{action}"; do_copy = False

            if do_copy:
                shutil.copy2(to_long_path(src_path), to_long_path(dest_path))
                if not self.simulate:
                    self._update_index(src_path, size, mtime, dest_path, md5_hex)
                if self.rename:
                    with self._count_lock:
                        self.count_renamed += 1

            self._append_row(src_path, dest_dir, new_name, cdt, extra, action, size, duplicate_of)
            self.status.emit(f"{action} ▶ {src_path.name} → {planned_dest}")
            self._update_counts(src_path)
        except Exception as e:
            self.status.emit(f"Error: {src_path.name} — {e}")

    def _update_counts(self, src_path: Path):
        ext = src_path.suffix.lower()
        with self._count_lock:
            if ext in IMAGE_EXTS: self.count_images += 1
            elif ext in VIDEO_EXTS: self.count_videos += 1

    def _append_row(
        self,
        src_path: Path,
        dest_dir: Path,
        new_name: str,
        cdt: datetime,
        extra: Dict[str, Optional[object]],
        action: str,
        size: int,
        duplicate_of: str,
    ):
        now_ts = datetime.now()
        row: Dict[str, object] = {
            "Timestamp": now_ts,
            "Source folder": str(src_path.parent),
            "Destination Folder": str(dest_dir),
            "Filename": src_path.name,
            "New Filename": new_name if self.rename else "",
            "Creation Date": cdt.date(),
            "Creation Time": cdt.time().replace(microsecond=0),
            "Action": action,
            "Size (bytes)": size,
            "Width": extra.get("width"),
            "Height": extra.get("height"),
            "Duration (sec)": extra.get("duration"),
            "Camera Make": extra.get("camera_make"),
            "Camera Model": extra.get("camera_model"),
            "Lens": extra.get("lens"),
            "GPS Lat": extra.get("gps_lat"),
            "GPS Lon": extra.get("gps_lon"),
            "Duplicate Of": duplicate_of,
        }
        with self._rows_lock:
            self.rows.append(row)

    # ---- Excel log ----
    def _write_excel_log(self, simulation: bool = False) -> Optional[Path]:
        wb = Workbook(); ws = wb.active; ws.title = "Log"
        headers = [
            "Timestamp","Source folder","Destination Folder","Filename","New Filename",
            "Creation Date","Creation Time","Action","Size (bytes)","Width","Height",
            "Duration (sec)","Camera Make","Camera Model","Lens","GPS Lat","GPS Lon","Duplicate Of",
        ]
        ws.append(headers)
        bold_font = Font(bold=True)
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_idx).font = bold_font

        for r in self.rows:
            ws.append([
                r.get("Timestamp"), r.get("Source folder"), r.get("Destination Folder"),
                r.get("Filename"), r.get("New Filename"), r.get("Creation Date"),
                r.get("Creation Time"), r.get("Action"), r.get("Size (bytes)"),
                r.get("Width"), r.get("Height"), r.get("Duration (sec)"),
                r.get("Camera Make"), r.get("Camera Model"), r.get("Lens"),
                r.get("GPS Lat"), r.get("GPS Lon"), r.get("Duplicate Of"),
            ])

        # Formats
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row: cell.number_format = "dd:mmmm:yyyy hh:mm:ss"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
            for cell in row: cell.number_format = "dd-mm-yyyy"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
            for cell in row: cell.number_format = "hh:mm:ss"

        # Auto width
        from openpyxl.utils import get_column_letter
        for col in range(1, len(headers) + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                try: max_len = max(max_len, len(str(val)))
                except Exception: pass
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)

        # One blank row, then bold summary
        summary_start = ws.max_row + 2
        ws.cell(row=summary_start, column=1, value="Summary").font = Font(bold=True)
        ws.cell(row=summary_start, column=2, value=f"Total files: {len(self.rows)}").font = Font(bold=True)
        ws.cell(row=summary_start, column=3, value=f"Images: {self.count_images}").font = Font(bold=True)
        ws.cell(row=summary_start, column=4, value=f"Videos: {self.count_videos}").font = Font(bold=True)
        ws.cell(row=summary_start, column=5, value=f"Renamed: {self.count_renamed}").font = Font(bold=True)
        ws.cell(row=summary_start, column=6, value=f"Workers: {self.workers_used}").font = Font(bold=True)
        ws.cell(row=summary_start, column=7, value=f"Mode: {'Simulate' if simulation else 'Copy'}").font = Font(bold=True)

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = "Simulation_Log_" if simulation else "Sort_Log_"
        log_name = f"{prefix}{stamp}.xlsx"
        log_path = self.dst_folder / log_name
        wb.save(log_path)
        return log_path

    # ---- index persistence ----
    def _load_index(self):
        try:
            if self._index_path.exists():
                with open(self._index_path, "r", encoding="utf-8") as f:
                    self._index = json.load(f)
        except Exception:
            self._index = {}

    def _save_index(self):
        try:
            with open(self._index_path, "w", encoding="utf-8") as f:
                json.dump(self._index, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

# ---------------------------
# UI
# ---------------------------
class PathPicker(QtWidgets.QWidget):
    changed = QtCore.Signal()
    def __init__(self, label: str, mode: str = "dir", parent=None):
        super().__init__(parent)
        self.mode = mode
        self.le = QtWidgets.QLineEdit()
        self.btn = QtWidgets.QPushButton("Browse…")
        self.btn.clicked.connect(self.browse)
        lay = QtWidgets.QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.addWidget(QtWidgets.QLabel(label))
        lay.addWidget(self.le, 1)
        lay.addWidget(self.btn)
        self.le.textChanged.connect(self.changed.emit)
    def text(self) -> str: return self.le.text().strip()
    def setText(self, t: str): self.le.setText(t)
    def browse(self):
        if self.mode == "dir":
            path = QtWidgets.QFileDialog.getExistingDirectory(self, "Select Folder")
        else:
            path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select File")
        if path: self.le.setText(path)

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.setMinimumSize(1150, 740)
        self.setWindowIcon(QtGui.QIcon())

        self.worker: Optional[SortWorker] = None
        self.current_log_path: Optional[str] = None

        central = QtWidgets.QWidget(); self.setCentralWidget(central)
        v = QtWidgets.QVBoxLayout(central); v.setContentsMargins(18, 18, 18, 18); v.setSpacing(14)

        # Theme toggle (keep default checker visuals)
        theme_bar = QtWidgets.QHBoxLayout(); theme_bar.addStretch(1)
        self.theme_toggle = QtWidgets.QCheckBox("Dark theme")
        self.theme_toggle.setChecked(True); self.theme_toggle.stateChanged.connect(self.apply_theme)
        theme_bar.addWidget(self.theme_toggle); v.addLayout(theme_bar)

        # Folders & Options
        box = QtWidgets.QGroupBox("Folders & Options")
        form = QtWidgets.QGridLayout(box); form.setHorizontalSpacing(12); form.setVerticalSpacing(14)

        self.src_picker = PathPicker("Source Folder:"); self.dst_picker = PathPicker("Destination Folder:")
        form.addWidget(self.src_picker, 0, 0, 1, 6); form.addWidget(self.dst_picker, 1, 0, 1, 6)

        # Checkboxes row (spaced)
        self.chk_recursive = QtWidgets.QCheckBox("Include subfolders (recursive)")
        self.chk_rename    = QtWidgets.QCheckBox("Append date to filename (DD-MM-YYYY-{original})")
        self.chk_resume    = QtWidgets.QCheckBox("Resume (skip already-copied)"); self.chk_resume.setChecked(True)
        self.chk_simulate  = QtWidgets.QCheckBox("Simulate (dry-run; no copying)")
        chk_row = QtWidgets.QHBoxLayout(); chk_row.setSpacing(22)
        chk_row.addWidget(self.chk_recursive); chk_row.addWidget(self.chk_rename)
        chk_row.addWidget(self.chk_resume);   chk_row.addWidget(self.chk_simulate); chk_row.addStretch(1)
        form.addLayout(chk_row, 2, 0, 1, 6)

        # Month language
        self.cmb_month_lang = QtWidgets.QComboBox(); self.cmb_month_lang.addItems(["English", "Indonesian"])
        form.addWidget(QtWidgets.QLabel("Month folder language:"), 3, 0); form.addWidget(self.cmb_month_lang, 3, 1)

        # Folder template (only 3 options)
        self.cmb_template = QtWidgets.QComboBox()
        self.cmb_template.addItem("YYYY/MonthName/DD", userData="ymd_name")
        self.cmb_template.addItem("YYYY/MM/DD", userData="ymd_mm")
        self.cmb_template.addItem("YYYY/Mon/DD", userData="ymd_mon")
        form.addWidget(QtWidgets.QLabel("Folder template:"), 3, 2); form.addWidget(self.cmb_template, 3, 3)

        # Duplicate detection + action
        self.cmb_dup_level = QtWidgets.QComboBox()
        self.cmb_dup_level.addItem("Off (default)", userData="off")
        self.cmb_dup_level.addItem("Exact (MD5) - detect duplicates", userData="exact")
        self.cmb_dup_action = QtWidgets.QComboBox()
        self.cmb_dup_action.addItem("Skip duplicates", userData="skip")
        self.cmb_dup_action.addItem("Keep both (rename if needed)", userData="keep")
        self.cmb_dup_action.addItem("Move to folder (inside date folder)", userData="folder")
        self.le_dup_folder = QtWidgets.QLineEdit("Duplicates")
        dup_row = QtWidgets.QHBoxLayout(); dup_row.setSpacing(12)
        dup_row.addWidget(QtWidgets.QLabel("Duplicate detection:")); dup_row.addWidget(self.cmb_dup_level)
        dup_row.addSpacing(16)
        dup_row.addWidget(QtWidgets.QLabel("When duplicate found:")); dup_row.addWidget(self.cmb_dup_action)
        dup_row.addSpacing(16)
        dup_row.addWidget(QtWidgets.QLabel("Folder name:")); dup_row.addWidget(self.le_dup_folder)
        dup_row.addStretch(1)
        form.addLayout(dup_row, 4, 0, 1, 6)

        v.addWidget(box)

        # Controls
        controls = QtWidgets.QHBoxLayout(); controls.addStretch(1)
        self.btn_preview   = QtWidgets.QPushButton("Preview (simulate)"); self.btn_preview.clicked.connect(self.start_preview)
        self.btn_start     = QtWidgets.QPushButton("Start"); self.btn_start.setMinimumHeight(42); self.btn_start.clicked.connect(self.start_work)
        self.btn_pause     = QtWidgets.QPushButton("Pause"); self.btn_pause.setEnabled(False); self.btn_pause.clicked.connect(self.toggle_pause)
        self.btn_stop      = QtWidgets.QPushButton("Stop");  self.btn_stop.setEnabled(False);  self.btn_stop.clicked.connect(self.stop_work)
        self.btn_clear_log = QtWidgets.QPushButton("Clear Log Screen"); self.btn_clear_log.clicked.connect(self.clear_log_screen)
        self.btn_open_log  = QtWidgets.QPushButton("Open Last Saved Log File"); self.btn_open_log.setEnabled(False); self.btn_open_log.clicked.connect(self.open_log)
        controls.addWidget(self.btn_preview); controls.addWidget(self.btn_start); controls.addWidget(self.btn_pause)
        controls.addWidget(self.btn_stop); controls.addWidget(self.btn_clear_log); controls.addWidget(self.btn_open_log)
        v.addLayout(controls)

        # Progress & Status
        self.progress = QtWidgets.QProgressBar(); self.progress.setValue(0)
        self.lbl_status = QtWidgets.QLabel("Ready."); self.lbl_status.setWordWrap(True)
        v.addWidget(self.progress); v.addWidget(self.lbl_status)

        # Log/Preview table
        self.table_preview = QtWidgets.QTableWidget(0, 5)
        self.table_preview.setHorizontalHeaderLabels(
            ["Filename", "Action", "Planned/New Name", "Destination Folder", "Duplicate Of"]
        )
        header = self.table_preview.horizontalHeader()
        header.setStretchLastSection(False)
        # Column sizing strategy for readability
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)  # Filename
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)  # Action
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)  # New Name
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.Stretch)           # Destination Folder (can be long)
        header.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeToContents)  # Duplicate Of
        self.table_preview.setWordWrap(False)
        self.table_preview.setTextElideMode(QtCore.Qt.ElideRight)
        self.table_preview.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table_preview.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        # Allow users to resize rows manually (default size kept)
        vheader = self.table_preview.verticalHeader()
        vheader.setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        vheader.setVisible(True)

        v.addWidget(self.table_preview)
        self.apply_theme()

    # Theme
    def apply_theme(self):
        dark = self.theme_toggle.isChecked()
        self.setStyleSheet(DARK_QSS if dark else LIGHT_QSS)

    # Helpers
    def clear_log_screen(self):
        self.table_preview.setRowCount(0)

    # Actions
    def _start_with_settings(self, simulate: bool):
        src = self.src_picker.text(); dst = self.dst_picker.text()
        if not src or not dst:
            QtWidgets.QMessageBox.warning(self, APP_TITLE, "Please choose both Source and Destination folders.")
            return False
        if Path(src) == Path(dst):
            QtWidgets.QMessageBox.warning(self, APP_TITLE, "Destination must be different from Source.")
            return False

        # Auto-clear log screen on each run
        self.clear_log_screen()

        self.progress.setValue(0)
        self.lbl_status.setText("Starting…")
        self.btn_start.setEnabled(False); self.btn_preview.setEnabled(False)
        self.btn_pause.setEnabled(True);  self.btn_stop.setEnabled(True)
        self.btn_open_log.setEnabled(False)

        lang = "id" if self.cmb_month_lang.currentText().lower().startswith("indo") else "en"
        template_key = self.cmb_template.currentData()
        dup_level = self.cmb_dup_level.currentData()
        dup_action = self.cmb_dup_action.currentData()
        dup_folder = self.le_dup_folder.text().strip() or "Duplicates"

        self.worker = SortWorker(
            src_folder=src,
            dst_folder=dst,
            recursive=self.chk_recursive.isChecked(),
            rename=self.chk_rename.isChecked(),
            month_lang=lang,
            template_key=template_key,
            simulate=simulate,
            resume_skip=self.chk_resume.isChecked(),
            dup_level=dup_level,
            dup_action=dup_action,
            dup_folder=dup_folder,
        )
        self.worker.progress.connect(self.on_progress)
        self.worker.status.connect(self.on_status)
        self.worker.error.connect(self.on_error)
        self.worker.finished_success.connect(self.on_finished)
        self.worker.start()
        return True

    def start_work(self):
        self._start_with_settings(simulate=self.chk_simulate.isChecked())

    def start_preview(self):
        self._start_with_settings(simulate=True)

    def toggle_pause(self):
        if not self.worker: return
        if self.btn_pause.text() == "Pause":
            self.worker.pause(); self.btn_pause.setText("Resume"); self.lbl_status.setText("Paused.")
        else:
            self.worker.resume(); self.btn_pause.setText("Pause"); self.lbl_status.setText("Resuming…")

    def stop_work(self):
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.lbl_status.setText("Stopping… (finishing current file)")

    def on_progress(self, processed: int, total: int):
        self.progress.setMaximum(max(total, 1)); self.progress.setValue(processed)

    def on_status(self, msg: str):
        self.lbl_status.setText(msg)

    def on_error(self, err: str):
        QtWidgets.QMessageBox.critical(self, APP_TITLE, f"Error: {err}")
        self.btn_start.setEnabled(True); self.btn_preview.setEnabled(True)
        self.btn_pause.setEnabled(False); self.btn_stop.setEnabled(False)

    def on_finished(self, summary: dict):
        self.btn_start.setEnabled(True); self.btn_preview.setEnabled(True)
        self.btn_pause.setEnabled(False); self.btn_stop.setEnabled(False)
        self.current_log_path = summary.get("log_path")
        if self.current_log_path and os.path.exists(self.current_log_path):
            self.btn_open_log.setEnabled(True)

        total = summary.get("total", 0); imgs = summary.get("images", 0)
        vids = summary.get("videos", 0); ren  = summary.get("renamed", 0)
        mode = "Simulate" if summary.get("simulate") else "Copy"
        msg = (
            f"Done [{mode}]. Processed: {total} | Images: {imgs} | Videos: {vids} | Renamed: {ren}.\n"
            f"Log saved to: {self.current_log_path or '(not created)'}"
        )
        self.lbl_status.setText(msg)

        if self.worker and hasattr(self.worker, "rows"):
            rows = self.worker.rows
            self.table_preview.setRowCount(len(rows))
            for i, r in enumerate(rows):
                vals = [
                    r.get("Filename", ""),
                    r.get("Action", ""),
                    r.get("New Filename", ""),
                    r.get("Destination Folder", ""),
                    r.get("Duplicate Of", ""),
                ]
                for j, val in enumerate(vals):
                    item = QtWidgets.QTableWidgetItem(str(val))
                    # Align columns nicely
                    if j == 1:  # Action
                        item.setTextAlignment(QtCore.Qt.AlignCenter)
                    else:
                        item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft)
                    # Long paths get tooltip
                    if j in (0, 2, 3, 4):
                        item.setToolTip(str(val))
                    self.table_preview.setItem(i, j, item)
            self.table_preview.resizeColumnsToContents()

    def open_log(self):
        if self.current_log_path and os.path.exists(self.current_log_path):
            QtGui.QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile(self.current_log_path))
        else:
            QtWidgets.QMessageBox.information(self, APP_TITLE, "No log file available.")

def main():
    app = QtWidgets.QApplication(sys.argv)
    app.setApplicationDisplayName(APP_TITLE)
    w = MainWindow(); w.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
